import inspect
import logging
from openpyxl import workbook, load_workbook
import softest as softest


class Utils(softest.TestCase):

    @staticmethod
    def get_logger():
        # create logger
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(logging.DEBUG)

        # Console Handler
        ch = logging.StreamHandler()
        fh = logging.FileHandler("logs/Automationlogs.log")

        # format
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s : %(message)s',
                                      datefmt='%m/%d/%Y %I:%M:%S %p')
        # add formatter to handler
        ch.setFormatter(formatter)
        fh.setFormatter(formatter)

        # add handler to logger
        logger.addHandler(ch)
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for r in range(2, row_ct + 1):
            row = []
            for c in range(1, col_ct + 1):
                row.append(sh.cell(row=r, column=c).value)
            datalist.append(row)

        return datalist

    def assert_equals(self, expected, actual):
        self.soft_assert(self.assertEqual, expected, actual)
        self.assert_all()